import openpyxl
from dataclasses import dataclass

ROW_AEM_LIVE = 2
ROW_EDITOR_URL = 3
ROW_SITE_CODE = 8
ROW_URL_TITLE = 13


@dataclass
class TranslationColumn:
    sheet_name: str
    col_idx: int
    site_code: str
    url_title: str
    editor_url: str


def load_workbook(path):
    return openpyxl.load_workbook(path)


def find_pending_columns(wb):
    """All sheets, all columns (B onward, A = row labels) where row 2 (AEM Live) is empty."""
    pending = []
    for sheet in wb.worksheets:
        for col in range(2, sheet.max_column + 1):
            live_cell = sheet.cell(row=ROW_AEM_LIVE, column=col)
            if live_cell.value not in (None, ""):
                continue
            site_code = sheet.cell(row=ROW_SITE_CODE, column=col).value
            url_title = sheet.cell(row=ROW_URL_TITLE, column=col).value
            editor_url = sheet.cell(row=ROW_EDITOR_URL, column=col).value
            if not site_code or not url_title:
                continue  # empty column, not a real translation
            pending.append(TranslationColumn(sheet.title, col, site_code, url_title, editor_url))
    return pending


def write_live_url(wb, path, sheet_name, col_idx, url):
    sheet = wb[sheet_name]
    sheet.cell(row=ROW_AEM_LIVE, column=col_idx, value=url)
    wb.save(path)  # save immediately, so progress survives a mid-run stop

