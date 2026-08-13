#!/usr/bin/env python3
import argparse
import json
import logging
import re
import subprocess
import sys
import unicodedata
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from urllib.parse import parse_qs, urlparse

logging.disable(logging.CRITICAL)
from openpyxl import load_workbook
from browser_owner import request as browser_request


ROOT = Path(__file__).resolve().parent
NODE_COPY = ROOT / "aem_sim_copy_child.mjs"
NODE_AUDIT = ROOT / "aem_faq_audit_page.mjs"
NODE_FIREFOX_FINALIZE = ROOT / "aem_firefox_finalize.mjs"
SHEET_HOSTS = {
    "Global": "https://p6spp-ap-author.samsung.com",
    "Europe": "https://p6spp-eu-author.samsung.com",
    "America": "https://p6spp-us-author.samsung.com",
}


def ui(event, **values):
    print("UI " + json.dumps({"kind": "qa", "event": event, **values}), flush=True)


def main():
    configure_output()
    args = parse_args()
    wb = load_workbook(args.workbook)
    if args.plan:
        print_plan(wb)
        return
    if args.retry_failed:
        retry_failed_copies(wb, args)
        return
    if args.all:
        run_all(wb, args)
        return

    ws = wb[args.sheet]
    target = find_column(ws, args.child_site)
    parent = parent_for_child(ws, target)
    host = host_for(ws, args.sheet)
    base = base_path(ws)
    source_path = f"/{ws.cell(8, parent).value}{base}{ws.cell(13, parent).value}/"
    destination_path = f"/{ws.cell(8, target).value}{base}{ws.cell(13, target).value}/"

    cmd = [
        "node",
        str(NODE_COPY),
        "--host", host,
        "--source-path", source_path,
        "--destination-path", destination_path,
        "--site-code", ws.cell(8, target).value,
    ]
    if args.detail_url:
        ids = detail_ids(args.detail_url)
        cmd += ["--content-id", ids["contentId"], "--request-id", ids["requestId"]]
    else:
        cmd += ["--slug", ws.cell(13, target).value]
    if args.apply:
        cmd.append("--apply")

    print(f"{'Copying' if args.apply else 'Dry run'} {args.sheet} {ws.cell(8, parent).value} -> {ws.cell(8, target).value}")
    subprocess.run(cmd, check=True)

    if args.apply:
        ws.cell(3, target).value = editor_url(host, ws.cell(8, target).value, base, ws.cell(13, target).value)
        save_workbook(wb, args.workbook)
        print(f"Wrote {ws.cell(3, target).coordinate}: {ws.cell(3, target).value}")


def parse_args():
    parser = argparse.ArgumentParser(description="QA/copy Samsung FAQ workbook columns.")
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--plan", action="store_true", help="print parent/child status from the workbook")
    parser.add_argument("--all", action="store_true", help="audit all parents; with --apply copy children of row-3-approved parents")
    parser.add_argument("--retry-failed", action="store_true", help="retry only approved children with no row-3 editor link")
    parser.add_argument("--review", action="store_true", help="pause for a row-3 approval decision after each unapproved parent audit")
    parser.add_argument("--copy-workers", type=int, default=3, help="concurrent child copies (default: 3)")
    parser.add_argument("--browser", choices=("firefox",), default="firefox")
    parser.add_argument("--user-data-dir", help="persistent Playwright profile for Firefox")
    parser.add_argument("--sheet", choices=sorted(SHEET_HOSTS), help="sheet containing the child")
    parser.add_argument("--child-site", help="AEM site code from row 8, e.g. africa_en")
    parser.add_argument("--detail-url", help="optional SIM detail URL fallback for an ambiguous inbox search")
    parser.add_argument("--apply", action="store_true", help="actually call SIM copy and write row 3")
    args = parser.parse_args()
    if not args.plan and not args.all and not args.retry_failed and not all([args.sheet, args.child_site]):
        parser.error("--sheet and --child-site are required unless --plan is used")
    return args


def configure_output():
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="backslashreplace")


def wait_for_workbook(path):
    while True:
        try:
            with open(path, "r+b"):
                return
        except PermissionError:
            print("WORKBOOK LOCKED: Close the .xlsx in Excel, then press Enter to retry.", flush=True)
            input()


def save_workbook(wb, path):
    while True:
        try:
            wb.save(path)
            return
        except PermissionError:
            print("WORKBOOK LOCKED: Close the .xlsx in Excel, then press Enter to retry.", flush=True)
            input()


def run_all(wb, args):
    browser = getattr(args, "browser", "firefox")
    user_data_dir = getattr(args, "user_data_dir", None)
    if (args.review or args.apply) and getattr(args, "workbook", None):
        wait_for_workbook(args.workbook)
    bridge = FirefoxBridge(user_data_dir) if browser == "firefox" else None
    if bridge:
        bridge.start()
    parents = []
    children = []
    for ws in wb.worksheets:
        yellow_parent = None
        for col in range(2, ws.max_column + 1):
            site = ws.cell(8, col).value
            if not site:
                continue
            kind = column_kind(ws.cell(11, col))
            if kind == "child":
                children.append((ws, col, yellow_parent))
            else:
                parents.append((ws, col))
                if kind == "parent-with-children":
                    yellow_parent = col

    baselines = {}
    findings = 0
    ui("start", parents=len(parents), children=len(children), child_sites=[ws.cell(8, col).value for ws, col, _ in children])
    for index, (ws, col) in enumerate(parents, 1):
        site = ws.cell(8, col).value
        link = editor_url(host_for(ws, ws.title), site, base_path(ws), ws.cell(13, col).value)
        ui("parent", index=index, total=len(parents), sheet=ws.title, site=site, link=link, status="auditing")
        print(f"PROGRESS parent {index}/{len(parents)} {ws.title}/{site}", flush=True)
        try:
            audit = audit_page(host_for(ws, ws.title), article_path(ws, col), link if args.review and not ws.cell(3, col).value else "", browser, user_data_dir, bridge)
        except (subprocess.CalledProcessError, subprocess.TimeoutExpired, RuntimeError) as error:
            findings += 1
            if isinstance(error, subprocess.TimeoutExpired):
                detail = f"audit timed out after {error.timeout}s"
            elif isinstance(error, RuntimeError):
                detail = str(error)
            else:
                lines = (error.stderr or error.stdout or str(error)).strip().splitlines()
                detail = lines[-1] if lines else str(error)
            print(f"FINDING {ws.title}/{site}: audit failed: {detail}", flush=True)
            ui("parent", index=index, total=len(parents), sheet=ws.title, site=site, link=link, status="error", findings=[detail])
            continue
        baseline = baselines.setdefault(ws.title, audit)
        page_findings = audit_findings(audit, baseline)
        for finding in page_findings:
            findings += 1
            print(f"FINDING {ws.title}/{site}: {finding}", flush=True)
        ui("parent", index=index, total=len(parents), sheet=ws.title, site=site, link=link, status="approved" if ws.cell(3, col).value else "pending", findings=page_findings)
        if args.review and not ws.cell(3, col).value:
            review_parent(wb, args.workbook, ws, col)

    copied = 0
    copy_jobs = {}
    workers = 1 if browser == "firefox" else max(1, args.copy_workers)
    with ThreadPoolExecutor(max_workers=workers) as executor:
        for index, (ws, child_col, parent_col) in enumerate(children, 1):
            child = ws.cell(8, child_col).value
            parent = ws.cell(8, parent_col).value if parent_col else "?"
            if not args.apply:
                print(f"COPY PLAN {index}/{len(children)} {ws.title}/{parent} -> {child}", flush=True)
                ui("child", index=index, total=len(children), sheet=ws.title, site=child, status="not copied")
                continue
            if not parent_col or not ws.cell(3, parent_col).value:
                print(f"COPY SKIP {ws.title}/{parent} -> {child}: parent is not row-3 approved", flush=True)
                ui("child", index=index, total=len(children), sheet=ws.title, site=child, status="not copied")
                continue
            print(f"PROGRESS copy {index}/{len(children)} {ws.title}/{parent} -> {child}", flush=True)
            ui("child", index=index, total=len(children), sheet=ws.title, site=child, status="copying")
            copy_jobs[executor.submit(copy_child, ws, ws.title, parent_col, child_col, browser, user_data_dir, bridge)] = (ws, child_col, parent, child)

        for future in as_completed(copy_jobs):
            ws, child_col, parent, child = copy_jobs[future]
            try:
                future.result()
            except (subprocess.CalledProcessError, RuntimeError) as error:
                print(f"COPY ERROR {ws.title}/{parent} -> {child}: {error}", flush=True)
                ui("child", sheet=ws.title, site=child, status="error", error=str(error))
                continue
            ws.cell(3, child_col).value = editor_url(host_for(ws, ws.title), child, base_path(ws), ws.cell(13, child_col).value)
            copied += 1
            save_workbook(wb, args.workbook)
            print(f"COPY DONE {ws.title}/{parent} -> {child}", flush=True)
            ui("child", sheet=ws.title, site=child, status="copied")
    if bridge:
        bridge.close()
    print(f"SUMMARY parents={len(parents)} findings={findings} children={len(children)} copied={copied}", flush=True)


def retry_failed_copies(wb, args):
    children = []
    for ws in wb.worksheets:
        yellow_parent = None
        for col in range(2, ws.max_column + 1):
            if not ws.cell(8, col).value:
                continue
            kind = column_kind(ws.cell(11, col))
            if kind == "parent-with-children":
                yellow_parent = col
            elif kind == "child" and yellow_parent and ws.cell(3, yellow_parent).value and not ws.cell(3, col).value:
                children.append((ws, col, yellow_parent))
    wait_for_workbook(args.workbook)
    bridge = FirefoxBridge(getattr(args, "user_data_dir", None))
    bridge.start()
    ui("start", parents=0, children=len(children), child_sites=[ws.cell(8, col).value for ws, col, _ in children])
    copied = 0
    for index, (ws, child_col, parent_col) in enumerate(children, 1):
        child, parent = ws.cell(8, child_col).value, ws.cell(8, parent_col).value
        print(f"PROGRESS retry {index}/{len(children)} {ws.title}/{parent} -> {child}", flush=True)
        ui("child", index=index, total=len(children), sheet=ws.title, site=child, status="copying")
        if child_exists(ws, ws.title, child_col, bridge):
            ws.cell(3, child_col).value = editor_url(host_for(ws, ws.title), child, base_path(ws), ws.cell(13, child_col).value)
            save_workbook(wb, args.workbook)
            print(f"COPY ALREADY DONE {ws.title}/{parent} -> {child}", flush=True)
            ui("child", sheet=ws.title, site=child, status="copied")
            continue
        try:
            copy_child(ws, ws.title, parent_col, child_col, "firefox", getattr(args, "user_data_dir", None), bridge)
        except (subprocess.CalledProcessError, RuntimeError) as error:
            print(f"COPY ERROR {ws.title}/{parent} -> {child}: {error}", flush=True)
            ui("child", sheet=ws.title, site=child, status="error", error=str(error))
            continue
        ws.cell(3, child_col).value = editor_url(host_for(ws, ws.title), child, base_path(ws), ws.cell(13, child_col).value)
        save_workbook(wb, args.workbook)
        copied += 1
        print(f"COPY DONE {ws.title}/{parent} -> {child}", flush=True)
        ui("child", sheet=ws.title, site=child, status="copied")
    bridge.close()
    print(f"SUMMARY retry children={len(children)} copied={copied}", flush=True)


def child_exists(ws, sheet, child_col, bridge):
    return bridge.request("exists", host=host_for(ws, sheet), path=article_path(ws, child_col))


def review_parent(wb, workbook_path, ws, col):
    site = ws.cell(8, col).value
    link = editor_url(host_for(ws, ws.title), site, base_path(ws), ws.cell(13, col).value)
    print(f"REVIEW ITEM {ws.title}/{site} {link}", flush=True)
    approved, note = review_answer(input("REVIEW RESPONSE y|n [note]: "))
    if approved:
        ws.cell(3, col).value = link
        save_workbook(wb, workbook_path)
        print(f"REVIEW APPROVED {ws.title}/{site}", flush=True)
        ui("parent", sheet=ws.title, site=site, link=link, status="approved", findings=[])
    else:
        print(f"REVIEW SKIPPED {ws.title}/{site}" + (f": {note}" if note else ""), flush=True)
        ui("parent", sheet=ws.title, site=site, link=link, status="not approved", findings=[note] if note else [])


def review_answer(value):
    choice, _, note = value.strip().partition(" ")
    return choice.lower() == "y", note.strip()


def audit_page(host, path, editor_link="", browser="firefox", user_data_dir=None, bridge=None):
    if bridge:
        return bridge.request("audit", host=host, path=path, editorUrl=editor_link)
    cmd = ["node", str(NODE_AUDIT), "--host", host, "--path", path]
    if editor_link:
        cmd += ["--editor-url", editor_link]
    if user_data_dir:
        cmd += ["--user-data-dir", user_data_dir]
    result = subprocess.run(
        cmd,
        check=True, capture_output=True, text=True, encoding="utf-8", timeout=60,
    )
    return json.loads(result.stdout)


def audit_findings(audit, baseline):
    components = audit["components"]
    expected = baseline["components"]
    findings = []
    for index, (component, reference) in enumerate(zip(components, expected), 1):
        if component["type"] == reference["type"] and component["settings"] != reference["settings"]:
            findings.append(f"component {index} settings differ from the sheet baseline")
        findings.extend(text_findings(component["text"]))
        if re.search(r"(?:^|/)co16(?:/|$)", component["type"], re.I) and any(re.search(r"<p\b", text, re.I) for text in component.get("descriptions", [])):
            findings.append(f"component {index} CO16 description contains a <p> tag")
    findings.extend(language_findings(components))
    return findings


def language_findings(components):
    """Flag a substantial unexpected writing system on one FAQ page."""
    texts = [text for component in components for text in component["text"]]
    scripts = [script(char) for text in texts for char in text]
    scripts = [name for name in scripts if name]
    if not scripts:
        return []
    expected = max(set(scripts), key=scripts.count)
    findings = []
    for text in texts:
        foreign = [name for name in (script(char) for char in text) if name and name not in (expected, "Latin")]
        if len(foreign) >= 3:
            findings.append(f"possible mixed language: {max(set(foreign), key=foreign.count)} text on a {expected} page")
            break
    return findings


def script(char):
    name = unicodedata.name(char, "")
    if "LATIN" in name:
        return "Latin"
    if "HANGUL" in name:
        return "Korean"
    if "HIRAGANA" in name or "KATAKANA" in name:
        return "Japanese"
    if "CJK" in name:
        return "Chinese"
    for marker, label in (
        ("ARABIC", "Arabic"), ("ARMENIAN", "Armenian"), ("BENGALI", "Bengali"),
        ("CYRILLIC", "Cyrillic"), ("DEVANAGARI", "Devanagari"), ("ETHIOPIC", "Ethiopic"),
        ("GEORGIAN", "Georgian"), ("GREEK", "Greek"), ("GUJARATI", "Gujarati"),
        ("GURMUKHI", "Gurmukhi"), ("HEBREW", "Hebrew"), ("KANNADA", "Kannada"),
        ("KHMER", "Khmer"), ("LAO", "Lao"), ("MALAYALAM", "Malayalam"),
        ("MYANMAR", "Burmese"), ("SINHALA", "Sinhala"), ("TAMIL", "Tamil"),
        ("TELUGU", "Telugu"), ("THAI", "Thai"), ("TIBETAN", "Tibetan"),
    ):
        if marker in name:
            return label
    return ""


def text_findings(values):
    text = "\n".join(values)
    findings = []
    if re.search(r"(?:\?\?|!!|\.\.)", text):
        findings.append("repeated punctuation")
    markers = re.findall(r"<(?:strong|b)\b[^>]*>\s*([^<]*\d+[^<]*)</(?:strong|b)>\s*([^\w\s<])?", text, re.I)
    labels = {step_label(label) for label, _ in markers}
    labels.discard("")
    if len(labels) > 1:
        findings.append("inconsistent bold numbered labels: " + ", ".join(sorted(labels)))
    punctuation = {}
    for label, mark in markers:
        label = step_label(label)
        if label:
            punctuation.setdefault(label, set()).add(mark or "none")
    for label, marks in punctuation.items():
        if len(marks) > 1:
            findings.append(f"differing punctuation after {label}: " + ", ".join(sorted(marks)))
    return findings


def step_label(value):
    return re.sub(r"\d+.*", "", re.sub(r"<[^>]+>", "", value)).strip().casefold()


def article_path(ws, col):
    return f"/content/samsung/{ws.cell(8, col).value}{base_path(ws)}{ws.cell(13, col).value}"


def copy_child(ws, sheet, parent, target, browser="firefox", user_data_dir=None, bridge=None):
    host = host_for(ws, sheet)
    source_path = article_path(ws, parent).removeprefix("/content/samsung") + "/"
    destination_path = article_path(ws, target).removeprefix("/content/samsung") + "/"
    if bridge:
        return bridge.request("copy", host=host, sourcePath=source_path, destinationPath=destination_path, siteCode=ws.cell(8, target).value, slug=ws.cell(13, target).value)
    cmd = [
        "node", str(NODE_COPY), "--host", host,
        "--source-path", source_path, "--destination-path", destination_path,
        "--site-code", ws.cell(8, target).value, "--slug", ws.cell(13, target).value,
    ]
    if user_data_dir:
        cmd += ["--user-data-dir", user_data_dir]
    cmd.append("--apply")
    subprocess.run(cmd, check=True)


class FirefoxBridge:
    def __init__(self, profile):
        self.profile = profile
        self.process = None

    def start(self):
        browser_request("done")

    def request(self, action, **values):
        return browser_request(action, **values)

    def close(self):
        pass

    def _read(self):
        line = self.process.stdout.readline()
        if not line:
            raise RuntimeError("Firefox session ended unexpectedly")
        return json.loads(line)


def print_plan(wb):
    for ws in wb.worksheets:
        yellow_parent = None
        parents = children = done = 0
        print(f"\n{ws.title}")
        for col in range(2, ws.max_column + 1):
            site = ws.cell(8, col).value
            if not site:
                continue
            kind = column_kind(ws.cell(11, col))
            if kind == "child":
                children += 1
                print(f"  child  {site:12} <- {ws.cell(8, yellow_parent).value if yellow_parent else '?'}")
            else:
                parents += 1
                if ws.cell(3, col).value:
                    done += 1
                if kind == "parent-with-children":
                    yellow_parent = col
        print(f"  parents: {parents}, QA marked: {done}, children: {children}")


def find_column(ws, site):
    matches = [col for col in range(2, ws.max_column + 1) if ws.cell(8, col).value == site]
    if len(matches) != 1:
        raise SystemExit(f"Expected one {site!r} column in {ws.title}, found {len(matches)}")
    if column_kind(ws.cell(11, matches[0])) != "child":
        raise SystemExit(f"{site} is not marked as a child column")
    return matches[0]


def parent_for_child(ws, child_col):
    for col in range(child_col - 1, 1, -1):
        if column_kind(ws.cell(11, col)) == "parent-with-children":
            return col
    raise SystemExit(f"No yellow parent before column {child_col}")


def column_kind(cell):
    color = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else None
    if color == "FFFCE5CD":
        return "child"
    if color == "FFFFFF00":
        return "parent-with-children"
    return "parent"


def host_for(ws, sheet):
    for col in range(2, ws.max_column + 1):
        url = ws.cell(3, col).value
        if url:
            parsed = urlparse(url)
            return f"{parsed.scheme}://{parsed.netloc}"
    return SHEET_HOSTS[sheet]


def base_path(ws):
    sample = ws.cell(6, 2).value or f"/{ws.cell(8, 2).value}/support/mobile-devices/{ws.cell(13, 2).value}"
    parts = sample.strip("/").split("/")
    return "/" + "/".join(parts[1:-1]) + "/"


def detail_ids(url):
    query = parse_qs(urlparse(url).query)
    content_id = first(query, "contentIdParam") or first(query, "contentId")
    request_id = first(query, "requestIdParam") or first(query, "requestId")
    if not content_id or not request_id:
        raise SystemExit("Detail URL must include contentIdParam/contentId and requestIdParam/requestId")
    return {"contentId": content_id, "requestId": request_id}


def first(query, key):
    values = query.get(key) or []
    return values[0] if values else ""


def editor_url(host, site, base, slug):
    return f"{host}/editor.html/content/samsung/{site}{base}{slug}.html"


if __name__ == "__main__":
    main()
