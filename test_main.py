import asyncio
import io
import unittest
from types import SimpleNamespace
from unittest.mock import AsyncMock, patch

import main
from main import LiveMonitor, skipped_countries
from openpyxl import Workbook
from sheet_io import find_columns


class LiveMonitorTest(unittest.TestCase):
    def test_releases_worker_slot(self):
        log = io.StringIO()
        monitor = LiveMonitor(2, log)
        slot = monitor.claim(SimpleNamespace(sheet_name="Europe", col_idx=22, site_code="se"))
        monitor.status(slot, "se: searching Jira")
        monitor.release(slot)
        self.assertEqual(monitor.slots, ["idle", "idle"])

    def test_live_check_saves_before_returning(self):
        class Monitor:
            def claim(self, col): return 0
            def release(self, slot): pass
            def status(self, slot, status): pass
            def log(self, message): pass

        col = SimpleNamespace(sheet_name="Europe", col_idx=22, site_code="se", editor_url="https://example.test/editor.html/content/samsung/se/page")
        with patch.object(main, "FILE_PATH", "test.xlsx", create=True), \
             patch.object(main, "check_live_async", new=AsyncMock(return_value=True)), \
             patch.object(main, "write_live_url") as write:
            result = asyncio.run(main._wait_for_live(object(), col, asyncio.Semaphore(1), Monitor(), object()))
        self.assertTrue(result[2])
        write.assert_called_once()

    def test_validate_all_includes_existing_live_urls(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Global"
        ws.cell(2, 2, "samsung.com/sg/page")
        ws.cell(3, 2, "https://example.test/editor.html/content/samsung/sg/page")
        ws.cell(8, 2, "sg")
        ws.cell(13, 2, "page")
        self.assertEqual([c.site_code for c in find_columns(wb, pending_only=False)], ["sg"])

    def test_skip_countries_accepts_commas_repeats_and_case(self):
        self.assertEqual(skipped_countries(["UK, ca", "SG"]), {"uk", "ca", "sg"})

    def test_publish_uses_a_fresh_firefox_context_with_the_signed_in_session(self):
        source = __import__("inspect").getsource(main.run_publish)
        self.assertIn('browser_request("storage_state")', source)
        self.assertIn("playwright.firefox.launch", source)
        self.assertIn("browser.new_context(storage_state=session)", source)

    def test_publish_checks_after_the_jira_phase(self):
        source = __import__("inspect").getsource(main.run_publish)
        self.assertIn("_wait_for_live", source)
        self.assertNotIn("_check_live(context, col, monitor, wb, slot)", __import__("inspect").getsource(main._publish_one))

    def test_publish_uses_v1_result_classification(self):
        source = __import__("inspect").getsource(main.run_publish)
        self.assertIn('result == "ambiguous"', source)

    def test_publish_serializes_jira_transitions(self):
        source = __import__("inspect").getsource(main.run_publish)
        self.assertIn("transition_lock = asyncio.Lock()", source)
        self.assertIn("_publish_one(context, col, semaphore, transition_lock, monitor)", source)

    def test_firefox_publish_defaults_to_one_worker(self):
        with patch("sys.argv", ["main.py", "--workbook", "workbook.xlsx"]):
            self.assertEqual(main.parse_args().workers, 1)

    def test_monitor_marks_skipped_countries_as_not_processed(self):
        monitor = LiveMonitor(1, io.StringIO())
        with patch("builtins.print") as print_:
            monitor.begin([SimpleNamespace(site_code="sg")], "publish", [SimpleNamespace(site_code="uk")])
        self.assertIn('"skipped": ["uk"]', print_.call_args.args[0])

    def test_validation_uses_parallel_shared_browser_requests(self):
        class Monitor:
            def __init__(self): self.slot = 0
            def claim(self, col): self.slot += 1; return self.slot
            def release(self, slot): pass
            def status(self, slot, status): pass

        active = peak = 0
        def check(action, url):
            nonlocal active, peak
            active += 1
            peak = max(peak, active)
            __import__("time").sleep(.01)
            active -= 1
            return True

        columns = [SimpleNamespace(sheet_name="Global", col_idx=i, site_code=str(i), editor_url=f"https://example.test/editor.html/content/samsung/{i}/page") for i in (2, 3)]
        with patch.object(main, "FILE_PATH", "test.xlsx", create=True), \
             patch.object(main, "browser_request", check), \
             patch.object(main, "write_live_url"):
            async def run():
                semaphore = asyncio.Semaphore(2)
                await asyncio.gather(*(main._validate_one(col, semaphore, Monitor(), object(), lambda _: None) for col in columns))
            asyncio.run(run())
        self.assertEqual(peak, 2)


if __name__ == "__main__":
    unittest.main()
