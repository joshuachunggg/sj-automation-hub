import argparse
import asyncio
import curses
import datetime
import sys
from openpyxl.utils import column_index_from_string
from playwright.async_api import async_playwright
from config import JIRA_BASE_URL, SESSION_FILE
from sheet_io import load_workbook, find_columns, find_pending_columns, write_live_url
from session_guard import dismiss_jira_notice, ensure_logged_in
from live_check import transform_editor_url, check_live, check_live_async
from jira_workflow import process_column


def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument("--workbook", required=True, help="Excel workbook to publish")
    p.add_argument("--sheet", help="only process this sheet name (test mode)")
    p.add_argument("--col", help="only process this column letter, e.g. AR (test mode)")
    p.add_argument("--start-sheet", help="resume run starting at this sheet")
    p.add_argument("--start-col", help="resume run starting at this column letter (needs --start-sheet)")
    p.add_argument("--workers", type=int, default=10, help="parallel Jira pages (1-15, default: 10)")
    p.add_argument("--browser", choices=("chromium", "firefox"), default="chromium", help="browser for Jira and live checks")
    p.add_argument(
        "--validate-only", action="store_true",
        help="skip Jira entirely - just re-check the live URL for pending columns and write back if live",
    )
    p.add_argument("--validate-all", action="store_true", help="skip Jira and check every locale URL, including columns already marked live")
    return p.parse_args()


class LiveMonitor:
    """Small terminal dashboard; the timestamped file remains the full append-only log."""

    def __init__(self, workers, log_file):
        self.workers = workers
        self.log_file = log_file
        self.slots = ["idle"] * workers
        self.events = []
        self.screen = None

    def start(self):
        if sys.stdout.isatty():
            self.screen = curses.initscr()
            curses.noecho()
            curses.cbreak()
            self.screen.nodelay(True)
            self.draw()

    def stop(self):
        if self.screen:
            curses.nocbreak()
            curses.echo()
            curses.endwin()
            self.screen = None

    def claim(self, col):
        slot = self.slots.index("idle")
        self.slots[slot] = f"{col.site_code}: starting"
        self.draw()
        return slot

    def release(self, slot):
        self.slots[slot] = "idle"
        self.draw()

    def status(self, slot, message):
        self.slots[slot] = message
        self.draw()

    def log(self, message):
        line = f"{datetime.datetime.now().strftime('%H:%M:%S')} {message}"
        self.log_file.write(line + "\n")
        self.events.append(line.replace("\n", " "))
        self.draw()
        if not self.screen:
            print(line)

    def draw(self):
        if not self.screen:
            return
        try:
            height, width = self.screen.getmaxyx()
            self.screen.erase()
            self.screen.addnstr(0, 0, f"AEM publisher  |  {sum(s != 'idle' for s in self.slots)}/{self.workers} workers active", width - 1)
            for index, state in enumerate(self.slots, 1):
                self.screen.addnstr(index, 0, f" {index:>2}. {state}", width - 1)
            log_start = self.workers + 2
            self.screen.addnstr(log_start, 0, "Live log", width - 1)
            for row, event in enumerate(self.events[-max(0, height - log_start - 1):], log_start + 1):
                self.screen.addnstr(row, 0, event, width - 1)
            self.screen.refresh()
        except curses.error:
            pass  # terminal is temporarily too small or being resized


def make_logger(workers):
    log_path = f"run_log_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
    log_file = open(log_path, "a", buffering=1)  # line-buffered, survives Ctrl+C
    monitor = LiveMonitor(workers, log_file)
    monitor.start()
    monitor.log(f"Logging this run to {log_path}")
    return monitor, log_file


def apply_start_from(pending, wb, start_sheet, start_col_letter):
    sheet_order = wb.sheetnames
    start_sheet_idx = sheet_order.index(start_sheet)  # raises if sheet name is wrong - fail loud
    start_col_idx = column_index_from_string(start_col_letter.upper())

    def keep(c):
        sheet_idx = sheet_order.index(c.sheet_name)
        if sheet_idx > start_sheet_idx:
            return True
        if sheet_idx == start_sheet_idx:
            return c.col_idx >= start_col_idx
        return False

    return [c for c in pending if keep(c)]


async def _validate_one(context, col, semaphore, monitor, wb, log):
    live_url = transform_editor_url(col.editor_url)
    async with semaphore:
        slot = monitor.claim(col)
        try:
            monitor.status(slot, f"{col.site_code}: checking live")
            if await check_live_async(context, live_url, timeout=5000):
                write_live_url(wb, FILE_PATH, col.sheet_name, col.col_idx, live_url)
                monitor.status(slot, f"{col.site_code}: live")
                log(f"[{col.sheet_name}] col {col.col_idx} ({col.site_code}): live: {live_url}")
                return None
            monitor.status(slot, f"{col.site_code}: not live")
            log(f"[{col.sheet_name}] col {col.col_idx} ({col.site_code}): not yet live: {live_url}")
            return col, live_url
        finally:
            monitor.release(slot)


async def run_validate_only(pending, wb, log, workers, monitor, browser_name):
    """No Jira or login: concurrently check locale URLs and save every confirmed result."""
    async with async_playwright() as p:
        browser = await getattr(p, browser_name).launch(headless=True)
        try:
            context = await browser.new_context()
            semaphore = asyncio.Semaphore(workers)
            results = await asyncio.gather(*(_validate_one(context, col, semaphore, monitor, wb, log) for col in pending))
        finally:
            await browser.close()
    still_not_live = [result for result in results if result]
    log("\n--- Validation Report ---")
    log(f"Still not live ({len(still_not_live)}): {[f'{c.sheet_name}/{c.site_code}' for c, _ in still_not_live]}")


async def _publish_one(context, col, semaphore, monitor):
    async with semaphore:
        slot = monitor.claim(col)
        try:
            for attempt in range(2):
                page = await context.new_page()
                try:
                    monitor.status(slot, f"{col.site_code}: opening Jira")
                    await page.goto(JIRA_BASE_URL, wait_until="domcontentloaded")
                    await dismiss_jira_notice(page)
                    await ensure_logged_in(page)
                    result = await process_column(
                        page, col, lambda message: monitor.status(slot, f"{col.site_code}: {message}"),
                    )
                    monitor.status(slot, f"{col.site_code}: {result}")
                    return col, result, None
                except Exception as e:
                    if attempt:
                        monitor.status(slot, f"{col.site_code}: error")
                        monitor.log(f"[{col.site_code}] [ERROR] {e}")
                        return col, None, str(e)
                    monitor.status(slot, f"{col.site_code}: retrying")
                    monitor.log(f"[{col.site_code}] retrying after: {e}")
                finally:
                    await page.close()
        finally:
            monitor.release(slot)


async def _wait_for_live(context, col, semaphore, monitor, wb):
    live_url = transform_editor_url(col.editor_url)
    async with semaphore:
        slot = monitor.claim(col)
        try:
            for attempt in range(4):
                monitor.status(slot, f"{col.site_code}: checking live ({attempt + 1}/4)")
                if await check_live_async(context, live_url):
                    write_live_url(wb, FILE_PATH, col.sheet_name, col.col_idx, live_url)
                    monitor.status(slot, f"{col.site_code}: live")
                    monitor.log(f"[{col.sheet_name}] col {col.col_idx} ({col.site_code}): verified live: {live_url}")
                    return col, live_url, True
                if attempt < 3:
                    await asyncio.sleep(15)
            monitor.status(slot, f"{col.site_code}: not live")
            monitor.log(f"[{col.sheet_name}] col {col.col_idx} ({col.site_code}): not live after 45s: {live_url}")
            return col, live_url, False
        finally:
            monitor.release(slot)


async def run_publish(pending, wb, log, workers, monitor, browser_name):
    not_found, ambiguous, errored = [], [], []
    async with async_playwright() as p:
        browser = await getattr(p, browser_name).launch(headless=False, slow_mo=100)
        context = await browser.new_context(storage_state=SESSION_FILE)
        try:
            probe = await context.new_page()
            await probe.goto(JIRA_BASE_URL, wait_until="domcontentloaded")
            await dismiss_jira_notice(probe)
            await ensure_logged_in(probe)
            await probe.close()

            semaphore = asyncio.Semaphore(workers)
            results = await asyncio.gather(*(_publish_one(context, col, semaphore, monitor) for col in pending))
            published = []
            for col, result, error in results:
                prefix = f"[{col.sheet_name}] col {col.col_idx}: {col.url_title} ({col.site_code})"
                if error:
                    log(f"{prefix}\n  [ERROR] {error}")
                    errored.append((col, error))
                elif result == "not_found":
                    log(f"{prefix}\n  not found")
                    not_found.append(col)
                elif result == "ambiguous":
                    log(f"{prefix}\n  ambiguous")
                    ambiguous.append(col)
                else:
                    published.append(col)

            checks = await asyncio.gather(*(_wait_for_live(context, col, semaphore, monitor, wb) for col in published))
            still_not_live = []
            for col, live_url, is_live in checks:
                if not is_live:
                    still_not_live.append((col, live_url))
        finally:
            await browser.close()

    log("\n--- Report ---")
    log(f"Not found ({len(not_found)}): {[c.url_title for c in not_found]}")
    log(f"Ambiguous ({len(ambiguous)}): {[c.url_title for c in ambiguous]}")
    log(f"Still not live ({len(still_not_live)}): {[f'{c.sheet_name}/{c.site_code}' for c, _ in still_not_live]}")
    log(f"Errored ({len(errored)}):")
    for c, err in errored:
        log(f"  {c.url_title} ({c.site_code}): {err}")
    not_confirmed = [*not_found, *ambiguous, *(c for c, _ in still_not_live), *(c for c, _ in errored)]
    log(f"Not confirmed live ({len(not_confirmed)}): {[f'{c.sheet_name}/{c.site_code}' for c in not_confirmed]}")


def main():
    args = parse_args()
    global FILE_PATH
    FILE_PATH = args.workbook
    if not 1 <= args.workers <= 15:
        raise SystemExit("--workers must be between 1 and 15")
    monitor, log_file = make_logger(args.workers)
    log = monitor.log
    try:
        wb = load_workbook(FILE_PATH)
        pending = find_columns(wb, pending_only=not args.validate_all)
        if args.sheet:
            pending = [c for c in pending if c.sheet_name == args.sheet]
        if args.col:
            col_idx = column_index_from_string(args.col.upper())
            pending = [c for c in pending if c.col_idx == col_idx]
        if args.start_sheet:
            pending = apply_start_from(pending, wb, args.start_sheet, args.start_col)
        log(f"{len(pending)} pending column(s) found across {len(wb.sheetnames)} sheet(s)")
        if args.validate_only or args.validate_all:
            asyncio.run(run_validate_only(pending, wb, log, args.workers, monitor, args.browser))
        else:
            asyncio.run(run_publish(pending, wb, log, args.workers, monitor, args.browser))
    finally:
        monitor.stop()
        log_file.close()


if __name__ == "__main__":
    main()
