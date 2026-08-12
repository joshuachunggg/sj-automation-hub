import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import hub
from openpyxl import Workbook


class DotenvTest(unittest.TestCase):
    def test_load_env_reads_values_without_exporting_them(self):
        with tempfile.TemporaryDirectory() as directory:
            env_file = Path(directory) / ".env"
            env_file.write_text("# local only\nWMC_USERNAME=user@example.com\nWMC_PASSWORD=secret\n")
            with patch.object(hub, "ENV_FILE", env_file):
                self.assertEqual(
                    hub.load_env(),
                    {"WMC_USERNAME": "user@example.com", "WMC_PASSWORD": "secret"},
                )

    def test_first_faq_editor_uses_the_first_translation_column(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "faq.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Global"
            ws.cell(6, 2, "/sg/support/mobile-devices/example")
            ws.cell(8, 2, "sg")
            ws.cell(13, 2, "example")
            wb.save(path)
            self.assertEqual(
                hub.first_faq_editor(path),
                "https://p6spp-ap-author.samsung.com/editor.html/content/samsung/sg/support/mobile-devices/example.html",
            )

    def test_macos_chrome_launch_uses_a_fresh_app_instance(self):
        chrome = "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"
        with patch.object(hub, "chrome_command", return_value=chrome), \
             patch.object(hub.platform, "system", return_value="Darwin"), \
             patch.object(hub.subprocess, "Popen") as popen:
            self.assertTrue(hub.launch_dev_chrome())
        self.assertEqual(popen.call_args.args[0][:3], ["open", "-na", "/Applications/Google Chrome.app"])

    def test_macos_chrome_activation_targets_chrome(self):
        with patch.object(hub.platform, "system", return_value="Darwin"), \
             patch.object(hub.subprocess, "run") as run:
            hub.activate_dev_chrome()
        self.assertEqual(run.call_args.args[0][:2], ["osascript", "-e"])


if __name__ == "__main__":
    unittest.main()
