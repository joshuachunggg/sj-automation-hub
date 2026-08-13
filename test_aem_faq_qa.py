import logging
import unittest
from types import SimpleNamespace
from unittest.mock import Mock, patch

logging.disable(logging.CRITICAL)
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from aem_faq_qa import FirefoxBridge, audit_findings, audit_page, column_kind, configure_output, detail_ids, language_findings, parent_for_child, parse_args, review_answer, review_parent, run_all, save_workbook, text_findings


class AemFaqQaTest(unittest.TestCase):
    def test_child_uses_previous_yellow_parent(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(8, 2).value = "sg"
        ws.cell(11, 2).fill = PatternFill("solid", fgColor="FFFFFF00")
        ws.cell(8, 3).value = "az"
        ws.cell(8, 4).value = "africa_en"
        ws.cell(11, 4).fill = PatternFill("solid", fgColor="FFFCE5CD")

        self.assertEqual(column_kind(ws.cell(11, 2)), "parent-with-children")
        self.assertEqual(column_kind(ws.cell(11, 3)), "parent")
        self.assertEqual(column_kind(ws.cell(11, 4)), "child")
        self.assertEqual(parent_for_child(ws, 4), 2)

    def test_detail_url_accepts_recorded_param_names(self):
        ids = detail_ids(
            "https://example.test/detail?siteCdParam=africa_en&contentIdParam=2204677&requestIdParam=631127"
        )
        self.assertEqual(ids, {"contentId": "2204677", "requestId": "631127"})

    def test_text_findings_detects_mixed_step_labels(self):
        findings = text_findings(["<p><strong>Step 1</strong></p><p><strong>Stage 2</strong></p>??"])
        self.assertIn("repeated punctuation", findings)
        self.assertIn("inconsistent bold numbered labels: stage, step", findings)

    def test_text_findings_detects_burmese_step_punctuation(self):
        findings = text_findings(["<b>အဆင့် ၁</b>. <b>အဆင့် ၂</b>။"])
        self.assertIn("differing punctuation after အဆင့်: ., ။", findings)

    def test_language_findings_flags_foreign_script(self):
        findings = language_findings([{"text": ["Restart your watch. အဆင့် သုံး"]}])
        self.assertEqual(findings, ["possible mixed language: Burmese text on a Latin page"])

    def test_language_findings_ignores_product_names_and_punctuation(self):
        self.assertEqual(language_findings([{"text": ["Restart Galaxy Watch 7."]}]), [])

    def test_language_findings_allows_english_on_any_page(self):
        self.assertEqual(language_findings([{"text": ["รีสตาร์ท Galaxy Watch เพื่อดำเนินการต่อ"]}]), [])

    def test_language_findings_flags_korean_on_chinese_page(self):
        findings = language_findings([{"text": ["請重新啟動手錶，然後檢查裝置是否可以正常運作。시계를 다시 시작하세요."]}])
        self.assertEqual(findings, ["possible mixed language: Korean text on a Chinese page"])

    def test_language_findings_flags_bengali_on_thai_page(self):
        findings = language_findings([{"text": ["รีสตาร์ทนาฬิกาแล้วตรวจสอบว่าทำงานได้ตามปกติ। ঘড়িটি পুনরায় চালু করুন।"]}])
        self.assertEqual(findings, ["possible mixed language: Bengali text on a Thai page"])

    def test_cli_output_uses_utf8(self):
        with patch("aem_faq_qa.sys.stdout") as stdout:
            configure_output()
        stdout.reconfigure.assert_called_once_with(encoding="utf-8", errors="backslashreplace")

    def test_audit_findings_ignores_component_count_and_order(self):
        audit = {"components": [{"type": "text", "settings": {}, "text": []}]}
        baseline = {"components": [
            {"type": "image", "settings": {}, "text": []},
            {"type": "text", "settings": {}, "text": []},
        ]}
        self.assertEqual(audit_findings(audit, baseline), [])

    def test_review_answer_keeps_a_skip_note(self):
        self.assertEqual(review_answer("y"), (True, ""))
        self.assertEqual(review_answer("n punctuation needs review"), (False, "punctuation needs review"))

    def test_copy_workers_default_to_three(self):
        with patch("sys.argv", ["aem_faq_qa.py", "--workbook", "workbook.xlsx", "--plan"]):
            self.assertEqual(parse_args().copy_workers, 3)

    def test_review_approval_writes_editor_link(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Global"
        ws.cell(6, 2).value = "/sg/support/mobile-devices/example"
        ws.cell(8, 2).value = "sg"
        ws.cell(13, 2).value = "example"
        with patch("builtins.input", return_value="y"), patch.object(wb, "save") as save:
            review_parent(wb, "workbook.xlsx", ws, 2)
        self.assertEqual(ws.cell(3, 2).value, "https://p6spp-ap-author.samsung.com/editor.html/content/samsung/sg/support/mobile-devices/example.html")
        save.assert_called_once_with("workbook.xlsx")

    def test_save_retries_after_workbook_unlock(self):
        workbook = Mock()
        workbook.save.side_effect = [PermissionError, None]
        with patch("builtins.input", return_value=""):
            save_workbook(workbook, "workbook.xlsx")
        self.assertEqual(workbook.save.call_count, 2)

    def test_audit_failure_without_stderr_does_not_stop_the_pass(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Global"
        ws.cell(8, 2).value = "sg"
        ws.cell(13, 2).value = "example"
        args = SimpleNamespace(apply=False, review=False, copy_workers=1, browser=None)
        error = __import__("subprocess").CalledProcessError(1, ["node"], stderr="")
        with patch("aem_faq_qa.audit_page", side_effect=error):
            run_all(wb, args)

    def test_audit_reads_node_output_as_utf8(self):
        result = SimpleNamespace(stdout='{"components": [{"text": ["အဆင့်"]}]}')
        with patch("aem_faq_qa.subprocess.run", return_value=result) as run:
            self.assertEqual(audit_page("https://example.test", "/content/example"), {"components": [{"text": ["အဆင့်"]}]})
        self.assertEqual(run.call_args.kwargs["encoding"], "utf-8")
        self.assertEqual(run.call_args.kwargs["timeout"], 60)

    def test_firefox_bridge_reads_utf8(self):
        with patch("aem_faq_qa.browser_request", return_value=True) as request:
            FirefoxBridge("profile").start()
        request.assert_called_once_with("done")

    def test_audit_timeout_does_not_stop_the_pass(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Global"
        ws.cell(8, 2).value = "sg"
        ws.cell(13, 2).value = "example"
        args = SimpleNamespace(apply=False, review=False, copy_workers=1, browser=None)
        with patch("aem_faq_qa.audit_page", side_effect=__import__("subprocess").TimeoutExpired(["node"], 60)):
            run_all(wb, args)

    def test_bad_firefox_editor_page_does_not_stop_the_pass(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Global"
        ws.cell(8, 2).value = "sg"
        ws.cell(13, 2).value = "example"
        args = SimpleNamespace(apply=False, review=True, copy_workers=1, browser=None)
        with patch("aem_faq_qa.audit_page", side_effect=RuntimeError("Could not open AEM editor: 404")):
            run_all(wb, args)


if __name__ == "__main__":
    unittest.main()
